import openpyxl
from openpyxl import Workbook
import os
import time

#defines
OFFSET_TO_START_FROM_ONE = 1 #starts from 0 but I want to start from 1
OFFSET_TO_CALC_NO_PPL = 1 #people in sheet actually start from 2nd pos
MIN_DAYS_TO_MATCH = 1 #minimum number of days that have to match to pair up people
MIN_SPORTS_TO_MATCH = 1 #minimum number of sports that have to match to pair up people

#define shortcuts for the answers
tak = 'Tak, chcę zostać z kimś połączony/a'
nie = 'Nie - dołączam z własnym Fitness Buddy'
wmale = 'Mężczyzną'
wfemale = 'Kobietą'
male = 'Mężczyzna'
female = 'Kobieta'
unknown = 'Inna'
both = 'Dowolnie'
online = 'Online'
station = 'Stacjonarnie'
preferDate = 'Preferowanego terminu ćwiczeń w tygodniu'
preferSport = 'Preferowanej dyscypliny sportu'
monday      = 0
tuesday     = 1
wednesday   = 2
thursday    = 3
friday      = 4
sat         = 5
sun         = 6

print("Starting...")

#specify the files location (or path)
cwd = os.path.dirname(os.path.abspath(__file__))
file = cwd + '\\fitness_buddy.xlsx'

#create an empty list to append values later on
values = []

#index of pairs with their buddy
pairIndex = 0

#global index of how many people were put into a sheet
numberOfPeopleInSheet = 0

#list of groups of people who look for a pair
listOfGroups = []
#list of pairs of people who already have a pair
hasPairGroup = []

#open file and worksheet
workbook = openpyxl.load_workbook(file)
worksheet = workbook.active
noOfEntries = worksheet.max_row - OFFSET_TO_CALC_NO_PPL

#create new workbook for assigned pairs
wbAssigned = Workbook()
wsAssigned =  wbAssigned.active

#create a class of people
class Person(object):
    def __init__(self, number):
        self.number = number
        self.freqz = ''
        self.onlineStationSet = ''
        self.sportSplit = ''
        self.days = [[monday],[tuesday],[wednesday],[thursday],[friday],[sat],[sun]]
        self.daysSplit = [[monday],[tuesday],[wednesday],[thursday],[friday],[sat],[sun]]
        self.genderBucket = 'notSpecified'

    #overload equals operator
    def __eq__(self, other):
        if (self.looksForBuddy == other.looksForBuddy and 
            self.genderBucket == other.genderBucket and
            self.freqz == other.freqz and
            self.onlineStationSet == other.onlineStationSet
            # and self.town == other.town #uncomment if want to add town constraint 
            ):
            return True
        else:
            return False

    #print person data
    def showData(self):
        print("number \t\t",        self.number + OFFSET_TO_START_FROM_ONE)
        print("name \t\t",          self.name)
        print("looksForBuddy \t",   self.looksForBuddy)
        print("nameToPair \t",      self.nameToPair)
        print("genderPair \t",      self.genderPair)
        print("onlineStation \t",   self.onlineStation)
        print("town \t",            self.town)
        print("dateSport \t",       self.dateSport)
        print("monday \t\t",        self.day[monday])
        print("tuesday \t",         self.day[tuesday])
        print("wednesday \t",       self.day[wednesday])
        print("thursday \t",        self.day[thursday])
        print("friday \t\t",        self.day[friday])
        print("sat \t\t",           self.day[sat])
        print("sun \t\t",           self.day[sun])
        print("discipline \t",      self.discipline)
        print("gender \t\t",        self.gender)
        print("freq \t\t",          self.freq)
        print("freqz \t\t",         self.freqz)
        print("genderBucket \t",    self.genderBucket)
        print(" ")

#create list of objects People
people = []
for i in range(noOfEntries):
    people.append(Person(i))    

#fill People's attributes with values from the worksheet
def initObject():
    for index, person in enumerate(people):
        i = index + 2 #index starts from zero so make i from 2 as the real rows start from 2nd
        person.name =           worksheet['B' + str(i)].value
        person.looksForBuddy =  worksheet['D' + str(i)].value
        person.nameToPair =     worksheet['E' + str(i)].value
        person.genderPair =     worksheet['F' + str(i)].value
        person.onlineStation =  worksheet['G' + str(i)].value
        person.town =           worksheet['H' + str(i)].value
        person.dateSport =      worksheet['I' + str(i)].value
        person.days[monday] =   worksheet['J' + str(i)].value
        person.days[tuesday] =  worksheet['K' + str(i)].value
        person.days[wednesday] =worksheet['L' + str(i)].value
        person.days[thursday] = worksheet['M' + str(i)].value
        person.days[friday] =   worksheet['N' + str(i)].value
        person.days[sat] =      worksheet['O' + str(i)].value
        person.days[sun] =      worksheet['P' + str(i)].value
        person.discipline =     worksheet['Q' + str(i)].value
        person.gender =         worksheet['R' + str(i)].value
        person.freq =           worksheet['S' + str(i)].value
        
initObject()

def openFile():
    try:
        os.startfile(cwd + '\\fitness_buddy_assigned.xlsx')
    except:
        print("error")

def forceCloseFile():
    try:
        os.system('TASKKILL /F /IM EXCEL.EXE')
        time.sleep(1)
    except:
        print("error")

#force shut excel file to be able to amend the assigned excel
forceCloseFile()

#setting all the variables for people
def setFrequency(person):
    if person.freq == 1 or person.freq == 2:
        person.freqz = 'low'
    elif person.freq == 3 or person.freq == 4:
        person.freqz = 'med'
    else:
        person.freqz = 'high'

def divideBySex(person):
    if person.gender == female and person.genderPair == wfemale:
        person.genderBucket = 'fwf'
    elif (person.gender == female and person.genderPair == wmale) or (person.gender == male and person.genderPair == wfemale):
        person.genderBucket = 'fwm'
    elif person.gender == male and person.genderPair == wmale:
        person.genderBucket = 'mwm'
    elif (person.gender == female and person.genderPair == both) or (person.gender == male and person.genderPair == both) or (person.gender == unknown and person.genderPair == both):
        person.genderBucket = 'both' #(f,m,u with both)
    elif (person.gender == unknown and person.genderPair == wfemale):
        person.genderBucket = 'uwf'
    elif (person.gender == unknown and person.genderPair == wmale):
        person.genderBucket = 'uwm'

#connect online with unspecified choice
def setOnlinePreference(person):
    if(person.onlineStation == online or person.onlineStation == both):
        person.onlineStationSet = online
    elif(person.onlineStation == station):
        person.onlineStationSet = station

def seperateDateSportPreference(person):
    if(person.dateSport == preferDate):
        for day in range(7): #bo 7 dni
            if(person.days[day] is not None): #some days aren't filled by the user so are empty, avoid them
                person.daysSplit[day] = person.days[day].split(sep=",") #seperate every comma and put values into a list
            else:
                person.daysSplit[day] = "None" #later I use it in appendPeopleToList() to filter out those cases
    elif(person.dateSport == preferSport):
        person.sportSplit = person.discipline.split(sep=",") #seperate every comma and put values into a list
        
def foundMatchingDays(person, others):
    timeMatchCount = 0
    dayMatchCount = 0
    for day in range(len(person.daysSplit)): #go through days
        if(person.daysSplit[day] is not None and others[0].daysSplit[day] is not None): #if day not empty then go in
            for time in range(len(person.daysSplit[day])): #go through time of the day for 1st person 
                for time2 in range(len(others[0].daysSplit[day])): #go through all times of the days for 2nd person 
                    if(person.daysSplit[day][time] == others[0].daysSplit[day][time2]): #if anything matches, count that in
                        if(person.daysSplit[day] == "None"): #here filtering out empty days
                            break
                        else:
                            timeMatchCount += 1
                            break
        if(timeMatchCount >= 1):
            dayMatchCount += 1 
            timeMatchCount = 0
    if(dayMatchCount >= MIN_DAYS_TO_MATCH):
        return True
    else:
        return False

def foundMatchingSports(person, others):
    sportMatchCount = 0
    for sport in range(len(person.sportSplit)):
        for sport2 in range(len(person.sportSplit)):
            if(person.sportSplit[sport] is not None and others[0].sportSplit[sport2] is not None): 
                if(person.sportSplit[sport] == others[0].sportSplit[sport2]):
                    sportMatchCount += 1
                    break
    if(sportMatchCount >= MIN_SPORTS_TO_MATCH):
        return True
    else:
        return False

#inserting people's all data into cells
def putPeopleInCell():
    global pairIndex
    global numberOfPeopleInSheet    
    for id in range(len(listOfGroups)):
        for idd in range(len(listOfGroups[id])):
            pairIndex+=1
            numberOfPeopleInSheet+=1
            wsAssigned['A' + str(pairIndex)] = numberOfPeopleInSheet
            # wsAssigned['A' + str(pairIndex)] = listOfGroups[id][idd].number
            wsAssigned['B' + str(pairIndex)] = listOfGroups[id][idd].name
            wsAssigned['C' + str(pairIndex)] = listOfGroups[id][idd].looksForBuddy 
            wsAssigned['D' + str(pairIndex)] = listOfGroups[id][idd].nameToPair 
            wsAssigned['E' + str(pairIndex)] = listOfGroups[id][idd].genderPair 
            wsAssigned['F' + str(pairIndex)] = listOfGroups[id][idd].onlineStation 
            wsAssigned['G' + str(pairIndex)] = listOfGroups[id][idd].town 
            wsAssigned['H' + str(pairIndex)] = listOfGroups[id][idd].dateSport
            wsAssigned['I' + str(pairIndex)] = listOfGroups[id][idd].days[monday]
            wsAssigned['J' + str(pairIndex)] = listOfGroups[id][idd].days[tuesday]
            wsAssigned['K' + str(pairIndex)] = listOfGroups[id][idd].days[wednesday]
            wsAssigned['L' + str(pairIndex)] = listOfGroups[id][idd].days[thursday]
            wsAssigned['M' + str(pairIndex)] = listOfGroups[id][idd].days[friday]
            wsAssigned['N' + str(pairIndex)] = listOfGroups[id][idd].days[sat]
            wsAssigned['O' + str(pairIndex)] = listOfGroups[id][idd].days[sun]
            wsAssigned['P' + str(pairIndex)] = listOfGroups[id][idd].discipline
            wsAssigned['Q' + str(pairIndex)] = listOfGroups[id][idd].gender 
            wsAssigned['R' + str(pairIndex)] = listOfGroups[id][idd].freq  

        pairIndex+=1

def putHasBuddyPeopleInCell():
    global pairIndex
    global numberOfPeopleInSheet      
    for id in range(len(hasPairGroup)):
        for idd in range(len(hasPairGroup[id])):
            pairIndex+=1
            numberOfPeopleInSheet+=1
            wsAssigned['A' + str(pairIndex)] = numberOfPeopleInSheet
            # wsAssigned['A' + str(pairIndex)] = hasPairGroup[id][idd].number
            wsAssigned['B' + str(pairIndex)] = hasPairGroup[id][idd].name
            wsAssigned['C' + str(pairIndex)] = hasPairGroup[id][idd].looksForBuddy 
            wsAssigned['D' + str(pairIndex)] = hasPairGroup[id][idd].nameToPair 
            wsAssigned['E' + str(pairIndex)] = hasPairGroup[id][idd].genderPair 
            wsAssigned['F' + str(pairIndex)] = hasPairGroup[id][idd].onlineStation 
            wsAssigned['G' + str(pairIndex)] = hasPairGroup[id][idd].town 
            wsAssigned['H' + str(pairIndex)] = hasPairGroup[id][idd].dateSport 
            wsAssigned['I' + str(pairIndex)] = hasPairGroup[id][idd].gender 
            wsAssigned['J' + str(pairIndex)] = hasPairGroup[id][idd].freq  
            wsAssigned['I' + str(pairIndex)] = hasPairGroup[id][idd].days[monday]
            wsAssigned['I' + str(pairIndex)] = hasPairGroup[id][idd].days[monday]
            wsAssigned['J' + str(pairIndex)] = hasPairGroup[id][idd].days[tuesday]
            wsAssigned['K' + str(pairIndex)] = hasPairGroup[id][idd].days[wednesday]
            wsAssigned['L' + str(pairIndex)] = hasPairGroup[id][idd].days[thursday]
            wsAssigned['M' + str(pairIndex)] = hasPairGroup[id][idd].days[friday]
            wsAssigned['N' + str(pairIndex)] = hasPairGroup[id][idd].days[sat]
            wsAssigned['O' + str(pairIndex)] = hasPairGroup[id][idd].days[sun]
            wsAssigned['P' + str(pairIndex)] = hasPairGroup[id][idd].discipline
            wsAssigned['Q' + str(pairIndex)] = hasPairGroup[id][idd].gender 
            wsAssigned['R' + str(pairIndex)] = hasPairGroup[id][idd].freq  
        pairIndex+=1

#adding people to specific groups
def addPersonHasPair(person):
    global hasPairGroup
    pairListWasExtended = False
    for grpIdx, groupElements in enumerate(hasPairGroup):
        if(person.nameToPair == groupElements[0].name):
            hasPairGroup[grpIdx].extend([person])
            pairListWasExtended = True
            break
    if(pairListWasExtended == False):
        hasPairGroup.append([person])

def appendPeopleToList(person):
    global listOfGroups
    listWasExtended = False
    for grpIdx, allElementsOfGroup in enumerate(listOfGroups): #go through all elements of listofgroups 
        if(person == allElementsOfGroup[0]): #take the first member of a current group and check if it matches person's preferences
            if(person.dateSport == preferDate): # if prefers by date then connect to other people with date preference
                if(foundMatchingDays(person, allElementsOfGroup)):
                    listOfGroups[grpIdx].extend([person]) #extend the current dimension of listofgroups (indicated by INDEX) by an element of current valueList element
                    listWasExtended = True
                    break
            elif(person.dateSport == preferSport):
                if(foundMatchingSports(person, allElementsOfGroup)):
                    listOfGroups[grpIdx].extend([person]) #extend the current dimension of listofgroups (indicated by INDEX) by an element of current valueList element
                    listWasExtended = True
                    break
        elif(person.looksForBuddy == nie):
            addPersonHasPair(person)
            listWasExtended = True
            break
    if(listWasExtended == False):
        listOfGroups.append([person]) #if no list was extended then create a new one at the end 

#main assigning loop                                     
def assignPeople():

    for person in people:
        #set person's frequency of excercises
        setFrequency(person)

        #rozdziel ludzi do kontenerow wzgledem plci i preferencji plci
        divideBySex(person)

        #set dowolnie preference to online
        setOnlinePreference(person)

        #set day preference or sport in one variable
        seperateDateSportPreference(person)

        #przydziel ludzi juz razem 
        appendPeopleToList(person)

#show all People's data
def showAllPeople():
    for person in people:
        person.showData()

#showAllPeople()

assignPeople()
putPeopleInCell()
putHasBuddyPeopleInCell()

#save and open new worksheet with added pairs
wbAssigned.save(cwd + '\\fitness_buddy_assigned.xlsx')
openFile()

#check if all people were successfully put into sheet
if(noOfEntries is not numberOfPeopleInSheet):
    print("MISSING PEOPLE!")
    print(noOfEntries - numberOfPeopleInSheet, "people were ommitted")
else:
    print("Success!")